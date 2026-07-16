"""
excel_report.py
Menghasilkan laporan Excel (.xlsx) untuk rapat dan tugas menggunakan openpyxl.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:

    HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin", color="E2E5EA"), right=Side(style="thin", color="E2E5EA"),
        top=Side(style="thin", color="E2E5EA"), bottom=Side(style="thin", color="E2E5EA")
    )

    def _style_header(self, ws, row=1):
        for cell in ws[row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _autofit(self, ws):
        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)

    def _apply_borders(self, ws, max_row, max_col):
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).border = self.THIN_BORDER

    def generate_meetings_report(self, filepath, meetings):
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Rapat"
        headers = ["No", "Judul Rapat", "Tanggal", "Waktu", "Lokasi", "Ketua Rapat", "Peserta", "Agenda"]
        ws.append(headers)
        for i, m in enumerate(meetings, start=1):
            ws.append([i, m.title, m.date, m.time, m.location, m.chairperson, m.participants, m.agenda])
        self._style_header(ws)
        self._apply_borders(ws, ws.max_row, ws.max_column)
        self._autofit(ws)
        wb.save(filepath)
        return filepath

    def generate_tasks_report(self, filepath, tasks):
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Tugas"
        headers = ["No", "Nama Tugas", "Penanggung Jawab", "Prioritas", "Deadline", "Status"]
        ws.append(headers)
        for i, t in enumerate(tasks, start=1):
            ws.append([i, t.task_name, t.assignee, t.priority, t.deadline, t.status])
        self._style_header(ws)
        self._apply_borders(ws, ws.max_row, ws.max_column)
        self._autofit(ws)
        wb.save(filepath)
        return filepath

    def generate_full_report(self, filepath, meetings, tasks, monthly_stats, completion_percentage):
        """Laporan gabungan dalam satu file dengan beberapa sheet."""
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Rapat"
        ws1.append(["No", "Judul Rapat", "Tanggal", "Waktu", "Lokasi", "Ketua Rapat"])
        for i, m in enumerate(meetings, start=1):
            ws1.append([i, m.title, m.date, m.time, m.location, m.chairperson])
        self._style_header(ws1)
        self._apply_borders(ws1, ws1.max_row, ws1.max_column)
        self._autofit(ws1)

        ws2 = wb.create_sheet("Tugas")
        ws2.append(["No", "Nama Tugas", "Penanggung Jawab", "Prioritas", "Deadline", "Status"])
        for i, t in enumerate(tasks, start=1):
            ws2.append([i, t.task_name, t.assignee, t.priority, t.deadline, t.status])
        self._style_header(ws2)
        self._apply_borders(ws2, ws2.max_row, ws2.max_column)
        self._autofit(ws2)

        ws3 = wb.create_sheet("Statistik Bulanan")
        ws3.append(["Bulan", "Jumlah Rapat"])
        for ym, total in monthly_stats:
            ws3.append([ym, total])
        self._style_header(ws3)
        self._apply_borders(ws3, ws3.max_row, ws3.max_column)
        self._autofit(ws3)

        ws4 = wb.create_sheet("Ringkasan")
        ws4.append(["Metrik", "Nilai"])
        ws4.append(["Total Rapat", len(meetings)])
        ws4.append(["Total Tugas", len(tasks)])
        ws4.append(["Persentase Penyelesaian Tugas", f"{completion_percentage}%"])
        self._style_header(ws4)
        self._apply_borders(ws4, ws4.max_row, ws4.max_column)
        self._autofit(ws4)

        wb.save(filepath)
        return filepath
